##link: https://www.youtube.com/watch?v=JUGEkFDeuwg
##execute this file in student_reg.py folder


import os
from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

##background color
background = "#06283D"
framebg="#EDEDED"
framefg="#06283D"

root=Tk()
root.title("Student Registration") ##tkinter heading
root.geometry("1250x700+80+10")  ##tkinter dimensions
root.config(bg=background)  ##tkinter bg colour

##create xlsx file
file=pathlib.Path('Student_data.xlsx')
if not os.path.isfile(file):
    
    wb = openpyxl.Workbook()
    #saving the workbook
    wb.save("Student_data.xlsx")

else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Regristration No"
    sheet['B1']="Name"
    sheet['C1']="Class"
    sheet['D1']="Gender"
    sheet['E1']="DOB"
    sheet['F1']="Date of Registration"
    sheet['G1']="Religion"
    sheet['H1']="Skill"
    sheet['I1']="Father Name"
    sheet['J1']="Mother Name"
    sheet['K1']="Father's Occupation"
    sheet['L1']="Mother's Occupation"

    file.save('Student_data.xlsx')

##select gender
def selection():
    global gender
    value=radio.get()
    if value==1:
        gender="Male"
        print(gender)
    else:
        gender="Female"
    
###############################################  exit window  ################################################
def Exit():
    root.destroy()

###########################################  select image from folder  ######################################
def showimage():
    
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),
                                        title="Select image file",filetype=(("JPG File","*.jpg"),
                                                                             ("PNG File","*.png"),
                                                                             ("All Files","*.txt")))
    
    img=Image.open(filename)
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2
         
###################################### registration no ##################################################
##it is created to automatic enter registration no
def registration_no():
    file=openpyxl.load_workbook('Student_data.xlsx')
    sheet=file.active
    row=sheet.max_row
    
    ##this will also check last registration no and add +1 if not present it will give 1
    max_row_value=sheet.cell(row=row,column=1).value
    print(max_row_value)
    try:
        Registration.set(max_row_value+1) ## append with registration number
    except:
        Registration.set("1") ##it will set to one if no registration no is present in xlsx file

########################################  clear #########################################
def Clear():
    global img
    Name.set('')
    DOB.set('')    
    Religion.set('')
    Skills.set('')    
    F_Name.set('')
    M_Name.set('')    
    Father_Occupation.set('')
    Mother_Occupation.set('')
    Class.set("Select Class")
    
    registration_no()
    
    saveButton.config(state='normal')
    
    # img1=PhotoImage(file="Image\upload_photo.png")
    img1=ImageTk.PhotoImage(Image.open(".\\Images\\upload_photo.png"))
    lbl.config(image=img1)
    lbl.iamge=img1
    
    img=""
    

#######################################  save  ############################################
def Save():
    R1=Registration.get()
    N1=Name.get()
    C1=Class.get()
    try:
        G1=gender  ##if not selected gender then it will show error
    except:
        messagebox.showerror("error","select gender")
    
    
    D2=DOB.get()
    D1=Date.get()
    Re1=Religion.get()
    S1=Skills.get()
    fathername=F_Name.get()
    mothername=M_Name.get()
    F1=Father_Occupation.get()
    M1=Mother_Occupation.get()
    
    
    ## error message if all details are filler
    if N1=="" or R1=="" or C1=="" or F1=="" or D2=="" or mothername=="" or fathername=="" or Re1=="" or D1=="":
        messagebox.showerror("error","details missing")
    
    ## save details to file  
    else:
        file=openpyxl.load_workbook("Student_data.xlsx")
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=R1)
        sheet.cell(column=2,row=sheet.max_row,value=N1)
        sheet.cell(column=3,row=sheet.max_row,value=C1)
        sheet.cell(column=4,row=sheet.max_row,value=G1)
        sheet.cell(column=5,row=sheet.max_row,value=D2)
        sheet.cell(column=6,row=sheet.max_row,value=D1)       
        sheet.cell(column=7,row=sheet.max_row,value=Re1)
        sheet.cell(column=8,row=sheet.max_row,value=S1)
        sheet.cell(column=9,row=sheet.max_row,value=fathername)
        sheet.cell(column=10,row=sheet.max_row,value=mothername)
        sheet.cell(column=11,row=sheet.max_row,value=F1)
        sheet.cell(column=12,row=sheet.max_row,value=M1)
        
        ##save to file
        file.save(r"Student_data.xlsx")
        
        ##save image with registration no
        try:
            img.save("Student Images/"+str(R1)+".jpg")
        except:
            messagebox.showinfo("info","picture not found")
        messagebox.showinfo("info","successfully data uploaded")  
        
        Clear()  ##clear entry box and iamge section
        
        registration_no()    ##it will recheck the registration no and reissue new no.        
    # print(D2)
    # print(D1)
    # print(Re1)
    # print(S1)
    # print(fathername)
    # print(mothername)
    # print(F1)
    # print(M1)   
    
######################################  search  ############################################
def search():
    
    text=Search.get()
    
    Clear() #to clear all data already available in entry box and other
    saveButton.config(state='disable') #after selecting on search, save button get disabled
    
    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active
    
    for row in sheet.rows:
        if row[0].value==int(text):
            name=row[0]
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]
    
    
    ##if registration is not present
    
    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid", "registration number is not valid")
        
        
    ##reg_no_position showing like A2,A3,A4,A5
    #but reg_number just showing number after A2 like 2,3...n
    
    x1=sheet.cell(row=int(reg_number),column=1).value
    x2=sheet.cell(row=int(reg_number),column=2).value
    x3=sheet.cell(row=int(reg_number),column=3).value
    x4=sheet.cell(row=int(reg_number),column=4).value
    x5=sheet.cell(row=int(reg_number),column=5).value
    x6=sheet.cell(row=int(reg_number),column=6).value
    x7=sheet.cell(row=int(reg_number),column=7).value
    x8=sheet.cell(row=int(reg_number),column=8).value
    x9=sheet.cell(row=int(reg_number),column=9).value
    x10=sheet.cell(row=int(reg_number),column=10).value
    x11=sheet.cell(row=int(reg_number),column=11).value
    x12=sheet.cell(row=int(reg_number),column=12).value
    
    # print(x1,x2,x3,x4,x5,x6,x7)  ## print details of already registration
    
    ##previous registration details in input box
    Registration.set(x1)
    Name.set(x2)
    Class.set(x3)
    if x4 =="Female":
        R2.select()
    else:
        R1.select()
    DOB.set(x5)
    Date.set(x6)
    Religion.set(x7)
    Skills.set(x8)
    F_Name.set(x9)
    M_Name.set(x10)
    Father_Occupation.set(x11)
    Mother_Occupation.set(x12)

    
    ##for image
    img=(Image.open("Student Images/"+str(x1)+".jpg"))
    resized_image=img.resize((190,190))
    photo2=ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2

#####################################   UPdate    ###########################################

def Update():
    R1=Registration.get()
    N1=Name.get()
    C1=Class.get()
    selection()
    G1=gender  
    D2=DOB.get()
    D1=Date.get()
    Re1=Religion.get()
    S1=Skills.get()
    fathername=F_Name.get()
    mothername=M_Name.get()
    F1=Father_Occupation.get()
    M1=Mother_Occupation.get()
    
    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active
    
    for row in sheet.rows:
        if row[0].value==R1:
            name=row[0]
            print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]
            
    # sheet.cell(column=1,row=int(reg_number),value=R1) ##non can update register number it will remain same
    sheet.cell(column=2,row=int(reg_number),value=N1)
    sheet.cell(column=3,row=int(reg_number),value=C1)
    sheet.cell(column=4,row=int(reg_number),value=G1)
    sheet.cell(column=5,row=int(reg_number),value=D2)
    sheet.cell(column=6,row=int(reg_number),value=D1)
    sheet.cell(column=7,row=int(reg_number),value=Re1)
    sheet.cell(column=8,row=int(reg_number),value=S1)
    sheet.cell(column=9,row=int(reg_number),value=fathername)
    sheet.cell(column=10,row=int(reg_number),value=mothername)
    sheet.cell(column=11,row=int(reg_number),value=F1)
    sheet.cell(column=12,row=int(reg_number),value=M1)

    file.save(r'Student_data.xlsx')
    
    try:
        img.save("Student Image/" + str(R1)+".jpg")
    except:
        pass
    messagebox.showinfo("Update","Update successful")

    Clear() ##clear after update
    
    
#####################################   top frame with email address    ###########################################
Label(root,text="Email Address:xyz123@gmail.com",width=10,height=3,bg="#f0687c",anchor='e').pack(side=TOP,fill=X)
Label(root,text="STUDENT REGISTRATION",width=10,height=2,bg="#c36464",fg="#fff",font='arial 20 bold').pack(side=TOP,fill=X)

##search box to update 
Search=StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font='arial 20 bold').place(x=820,y=70)
imageicon3=ImageTk.PhotoImage(Image.open(".\\Images\\search.png"))
# imageicon3=PhotoImage(file="D:\\learning_python\\py_files\\task_file\\student_registrationwith_database\\Images\\search.png")
Srch=Button(root,text="Search",compound=LEFT,image=imageicon3,width=123,height=35,bg='#68ddfa',font="arial 13 bold",command=search)
Srch.place(x=1060,y=66)  ##set dimension where to place(width, height)


imageicon4=ImageTk.PhotoImage(Image.open(".\\Images\\Layer.png"))
Update_button=Button(root,image=imageicon4,width=60,height=35,bg="#c36464",command=Update)
Update_button.place(x=110,y=64)  ##set dimension where to place


###################################### registration and date ###############################################
Label(root,text="Regristration No:",font="arial 13",fg=framebg,bg=background).place(x=30,y=150)
Label(root,text="Date:",font="arial 13",fg=framebg,bg=background).place(x=500,y=150)

Registration=IntVar()   ##int variable for registration no
Date=StringVar()   ##string variable for date variable

##################################### display registration input box #######################################
reg_entry=Entry(root,textvariable=Registration,width=15,font="arial 10")
reg_entry.place(x=160,y=150)  ##set dimension where to place

registration_no()  ##called registration_no() function


today=date.today()
d1=today.strftime("%d-%m-%y")  ##current date

################################  display date input box ###################################################
date_entry=Entry(root,textvariable=Date,width=15,font="arial 10")
date_entry.place(x=550,y=150)    ##set dimension where to place

Date.set(d1) ##set currentdate

###################################### student details  ################################################### 
obj=LabelFrame(root,text="Student's Details",font=20,bd=20,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)   ##set dimension where to place(width, height)

Label(obj,text="Full Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj,text="Date Of Birth:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)


Label(obj,text="Class:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj,text="Religion:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(obj,text="Skills:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)

#################################### input box for student deatils  ##############################################
Name=StringVar() ##string variable for name variable
name_entry=Entry(obj,textvariable=Name,width=20,font="arial 10")  ##box size and font of input
name_entry.place(x=160,y=50) ##set dimension where to place(width, height)

DOB=StringVar() ##string variable for DOB variable
dob_entry=Entry(obj,textvariable=DOB,width=20,font="arial 10")  ##box size and font of input
dob_entry.place(x=160,y=100) ##set dimension where to place(width, height)


#gender radio button
radio=IntVar()   ##int variable for gender variable

##male radio button
R1=Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=150,y=150)

##female radio button
R2=Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
R2.place(x=200,y=150)

##regilion input box
Religion=StringVar() ##string variable for religion variable
religion_entry=Entry(obj,textvariable=Religion,width=20,font="arial 10")  ##box size and font of input
religion_entry.place(x=630,y=100) ##set dimension where to place(width, height)

##skills input box
Skills=StringVar() ##string variable for skils variable
skills_entry=Entry(obj,textvariable=Skills,width=20,font="arial 10")  ##box size and font of input
skills_entry.place(x=630,y=150) ##set dimension where to place(width, height)

##class with options
Class=Combobox(obj,values=["1","2","3","4","5","6","7","8","9","10"],font="Robot 13",width=17,state="r")
Class.place(x=630,y=50)
Class.set("Select Class")  ##display name for class

##parents details   
obj2=LabelFrame(root,text="Parent's Details",font=20,bd=20,width=900,bg=framebg,fg=framefg,height=220,relief=GROOVE)
obj2.place(x=30,y=470)   ##set dimension where to place(width, height)

##father name as key word
Label(obj2,text="Father's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)

##input box for father name and occupation
F_Name=StringVar()
f_entry=Entry(obj2,textvariable=F_Name,width=20,font="arial 10")
f_entry.place(x=160,y=50)

Father_Occupation=StringVar()
fo_entry=Entry(obj2,textvariable=Father_Occupation,width=20,font="arial 10")
fo_entry.place(x=160,y=100)

##mother name as key word
Label(obj2,text="Mother's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)

##input box for father name and occupation
M_Name=StringVar()
m_entry=Entry(obj2,textvariable=M_Name,width=20,font="arial 10")
m_entry.place(x=630,y=50)

Mother_Occupation=StringVar()
mo_entry=Entry(obj2,textvariable=Mother_Occupation,width=20,font="arial 10")
mo_entry.place(x=630,y=100)

##image
f=Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)

img=ImageTk.PhotoImage(Image.open(".\\Images\\upload_photo.png"))
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)

##create button  save,upload,reset,exit
Button(root,text="Upload",width=19,height=2,font="arial 12 bold",bg="lightblue",command=showimage).place(x=1000,y=370)
saveButton=Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="lightblue",command=Save)
saveButton.place(x=1000,y=450)
Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="lightblue",command=Clear).place(x=1000,y=530)
Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="lightblue",command=Exit).place(x=1000,y=610)


root.mainloop()
